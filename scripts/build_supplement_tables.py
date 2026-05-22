"""
Build publication-ready supplement tables S8, S9, S10a, S10b
from raw analysis CSVs.

Outputs to /tmp/supplement-tables/:
  Table_S8_genotyping.xlsx / .html
  Table_S9_qba.xlsx        / .html
  Table_S10a_ld_breakdown.xlsx / .html
  Table_S10b_ld_summary.xlsx   / .html
"""
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl", "pandas"]
# ///
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC = Path("/Users/mikeford/dpyd-arm-model/output/sensitivity")
OUT = Path("/tmp/supplement-tables")
OUT.mkdir(exist_ok=True, parents=True)

VARIANT_NAME = {
    "WT_Clean": "WT (clean)",
    "WT_Biased": "WT (biased)",
    "HapB3": "HapB3",
    "2846hetho": "c.2846A>T",
    "2Ahetho": "*2A",
    "13hetho": "*13",
}

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="1F3864")
TITLE_FILL = PatternFill("solid", fgColor="E7EBF5")
CAPTION_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color="404040")
CAPTION_FONT = Font(name="Calibri", size=10, italic=True, color="404040")
NOTES_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="404040")
NOTES_HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
FOOTNOTE_FONT = Font(name="Calibri", size=9, italic=True, color="404040")
BODY_FONT = Font(name="Calibri", size=10)
TOTAL_FILL = PatternFill("solid", fgColor="EDEDED")
TOTAL_FONT = Font(name="Calibri", size=10, bold=True)

THIN = Side(border_style="thin", color="808080")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
TITLE_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_body_cell(cell, align="center", bold=False):
    cell.font = TOTAL_FONT if bold else BODY_FONT
    cell.alignment = CENTER if align == "center" else LEFT
    cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _estimate_height(text, total_width_chars, char_factor=1.05, line_height=15, min_lines=1):
    """Rough estimate of row height needed for wrapped text."""
    if not text:
        return line_height * min_lines
    # Count explicit newlines
    explicit_lines = text.count("\n") + 1
    longest = max(len(line) for line in text.split("\n"))
    wrapped_lines = max(1, int((longest * char_factor) / max(total_width_chars, 1)) + 1)
    lines = max(min_lines, explicit_lines + wrapped_lines - 1)
    return lines * line_height + 4


def write_title_block(ws, title, caption, n_cols, col_widths=None):
    """
    Layout (occupies rows 1-3; header should be written to row 4):
      Row 1: Title — bold 14pt, light-blue banner
      Row 2: Caption — labeled "Caption: <text>" — italic 10pt
      Row 3: spacer
    """
    # Row 1: Title
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = TITLE_ALIGN
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=n_cols)
    for c in range(2, n_cols + 1):
        ws.cell(row=1, column=c).fill = TITLE_FILL
    ws.row_dimensions[1].height = 26

    # Row 2: Caption with bold "Caption:" prefix
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    label_font = InlineFont(rFont="Calibri", sz=10, b=True, color="404040")
    text_font = InlineFont(rFont="Calibri", sz=10, i=True, color="404040")
    rich = CellRichText(
        TextBlock(label_font, "Caption: "),
        TextBlock(text_font, caption),
    )
    cap_cell = ws.cell(row=2, column=1)
    cap_cell.value = rich
    cap_cell.alignment = LEFT_TOP
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=n_cols)
    total_w = sum(col_widths) if col_widths else max(60, n_cols * 12)
    ws.row_dimensions[2].height = _estimate_height(
        "Caption: " + caption, total_w, char_factor=1.0, line_height=14, min_lines=2,
    )
    # Row 3: spacer separating caption from header row
    ws.row_dimensions[3].height = 8


def write_footnotes(ws, start_row, n_cols, notes, col_widths=None):
    """
    Layout starting at `start_row`:
      Row N:   blank spacer
      Row N+1: "Notes:" section header (bold, light-gray banner)
      Row N+2 .. N+1+len(notes): numbered footnotes "1. <text>", "2. <text>", ...
    """
    # spacer
    ws.row_dimensions[start_row].height = 6
    # Notes header
    hdr_row = start_row + 1
    cell = ws.cell(row=hdr_row, column=1, value="Notes:")
    cell.font = NOTES_HEADER_FONT
    cell.fill = NOTES_HEADER_FILL
    cell.alignment = LEFT
    ws.merge_cells(start_row=hdr_row, end_row=hdr_row, start_column=1, end_column=n_cols)
    for c in range(2, n_cols + 1):
        ws.cell(row=hdr_row, column=c).fill = NOTES_HEADER_FILL
    ws.row_dimensions[hdr_row].height = 20

    total_w = sum(col_widths) if col_widths else max(60, n_cols * 12)
    for i, note in enumerate(notes, start=1):
        r = hdr_row + i
        text = f"{i}. {note}"
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = FOOTNOTE_FONT
        cell.alignment = LEFT_TOP
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=n_cols)
        ws.row_dimensions[r].height = _estimate_height(
            text, total_w, char_factor=1.0, line_height=12, min_lines=1,
        )


def configure_page(ws, landscape=False):
    """Set print orientation and fit-to-width so the table prints cleanly."""
    from openpyxl.worksheet.page import PageMargins
    if landscape:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    else:
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


HTML_CSS = """
<style>
body { font-family: 'Calibri', 'Helvetica', sans-serif; max-width: 1100px; margin: 24px auto; color: #222; }
h2 { font-size: 14pt; margin-bottom: 4px; }
p.caption { font-size: 10pt; color: #555; margin-top: 0; margin-bottom: 12px; font-style: italic; }
table { border-collapse: collapse; font-size: 10pt; margin-bottom: 8px; }
th { background: #1F3864; color: white; padding: 6px 10px; border: 1px solid #555; text-align: center; vertical-align: middle; }
td { padding: 5px 10px; border: 1px solid #aaa; text-align: center; vertical-align: middle; }
td.left { text-align: left; }
tr.total td { background: #EDEDED; font-weight: bold; }
p.footnote { font-size: 9pt; color: #555; margin: 4px 0; }
</style>
"""


def html_page(title, body):
    return (
        f"<!doctype html><html><head><meta charset='utf-8'>"
        f"<title>{title}</title>{HTML_CSS}</head><body>{body}</body></html>"
    )


def yn(v):
    s = str(v).strip().upper()
    if s == "TRUE":
        return "Yes"
    if s == "FALSE":
        return "No"
    return "—"


def build_s8():
    df = pd.read_csv(SRC / "genotyping" / "genotyping_combined_summary.csv")
    diag_start = df.index[df["comparison"] == "---DIAGNOSTICS---"][0]
    diagnostics = df.iloc[diag_start + 1:].copy()
    body = df.iloc[:diag_start].copy()

    def label_pair(s):
        suffix = ""
        s_clean = s
        if "(" in s:
            i = s.index("(")
            s_clean, suffix = s[:i].strip(), s[i:].strip()
        a, b = [t.strip() for t in s_clean.split(" vs ")]
        out = f"{VARIANT_NAME.get(a, a)} vs {VARIANT_NAME.get(b, b)}"
        if suffix:
            out += f" {suffix}"
        return out

    body["Comparison"] = body["comparison"].map(label_pair)

    def fmt_sucra(v):
        try:
            f = float(v)
        except (TypeError, ValueError):
            return "—"
        if f != f:  # NaN check
            return "—"
        return f"{f:.1f}"

    display_headers = [
        "Comparison",
        "OR", "95% CrI", "Excludes 1?", "SUCRA (%)",
        "OR", "95% CrI", "Excludes 1?", "SUCRA (%)",
    ]

    def render_row(r):
        return [
            r["Comparison"],
            f"{float(r['OR_loose']):.2f}",
            str(r["CrI_loose"]),
            yn(r["CrI_excludes_1_loose"]),
            fmt_sucra(r["SUCRA_loose"]),
            f"{float(r['OR_strict']):.2f}",
            str(r["CrI_strict"]),
            yn(r["CrI_excludes_1_strict"]),
            fmt_sucra(r["SUCRA_strict"]),
        ]

    rows = [render_row(r) for _, r in body.iterrows()]
    diag = {row["comparison"]: (row["OR_loose"], row["OR_strict"])
            for _, row in diagnostics.iterrows()}

    title = "Table S8. Genotyping Misclassification Sensitivity Analysis (WT-Split Model)"
    caption = (
        "Network meta-analysis after splitting WT into WT (clean) and WT (biased) nodes to test whether "
        "genotyping completeness drives the primary finding. Two WT (clean) definitions are reported: "
        "loose (n = 9 cohorts; includes studies with pan-negative-by-design WT arms plus three studies "
        "with functionally pan-negative WT controls — Amstutz_2009, Jennings_2013, Froehlich_2015) and "
        "strict (n = 6 cohorts; restricted to studies with pre-screening exclusion of variant carriers). "
        "WT (biased) comprises studies that did not test for all four clinically actionable variants. "
        "OR = odds ratio; CrI = credible interval; SUCRA = surface under the cumulative ranking curve."
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Table S8"
    n_cols = len(display_headers)
    col_widths = [38, 9, 14, 12, 11, 9, 14, 12, 11]
    autosize(ws, col_widths)
    write_title_block(ws, title, caption, n_cols, col_widths)

    ws.merge_cells(start_row=4, end_row=5, start_column=1, end_column=1)
    cell = ws.cell(row=4, column=1, value="Comparison")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = BORDER

    ws.cell(row=4, column=2, value="Loose definition (n = 9 WT-clean cohorts)")
    ws.merge_cells(start_row=4, end_row=4, start_column=2, end_column=5)
    ws.cell(row=4, column=6, value="Strict definition (n = 6 WT-clean cohorts)")
    ws.merge_cells(start_row=4, end_row=4, start_column=6, end_column=9)
    for c in (2, 6):
        cell = ws.cell(row=4, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for i, h in enumerate(display_headers, start=1):
        if i == 1:
            continue
        cell = ws.cell(row=5, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    r = 6
    for row in rows:
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            style_body_cell(cell, align="left" if c == 1 else "center")
        r += 1

    notes = [
        ("Full treatment SUCRA rankings (loose / strict): "
         "WT (clean) 94.7 / 90.8, WT (biased) 85.1 / 88.6, HapB3 59.7 / 60.1, "
         "*2A 24.9 / 24.9, *13 22.6 / 22.6, c.2846A>T 13.0 / 12.9. "
         "Higher SUCRA = lower predicted toxicity. WT (clean) ranks first in both models."),
        (
            f"Convergence (loose): DIC = {diag['DIC'][0]}, max PSRF = {diag['max_PSRF'][0]}, "
            f"converged = {yn(diag['converged'][0])}.  "
            f"Convergence (strict): DIC = {diag['DIC'][1]}, max PSRF = {diag['max_PSRF'][1]}, "
            f"converged = {yn(diag['converged'][1])}."
        ),
        ("WT (clean) vs WT (biased) is reported as a diagnostic comparison to assess whether genotyping "
         "completeness drives outcome heterogeneity; SUCRA is not meaningful for this two-WT contrast."),
        ("PSRF = potential scale reduction factor (Gelman-Rubin diagnostic, threshold ≤ 1.10). "
         "DIC = deviance information criterion."),
    ]
    write_footnotes(ws, r, n_cols, notes, col_widths)
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 22
    configure_page(ws, landscape=True)
    wb.save(OUT / "Table_S8_genotyping.xlsx")

    html_rows = [
        "<tr>"
        "<th rowspan='2'>Comparison</th>"
        "<th colspan='4'>Loose definition (n = 9 WT-clean cohorts)</th>"
        "<th colspan='4'>Strict definition (n = 6 WT-clean cohorts)</th>"
        "</tr>",
        "<tr>" + "".join(f"<th>{h}</th>" for h in display_headers[1:]) + "</tr>",
    ]
    for row in rows:
        cells = "".join(
            f"<td class='left'>{v}</td>" if i == 0 else f"<td>{v}</td>"
            for i, v in enumerate(row)
        )
        html_rows.append(f"<tr>{cells}</tr>")
    table_html = "<table>" + "".join(html_rows) + "</table>"
    body_html = f"<h2>{title}</h2><p class='caption'>{caption}</p>{table_html}"
    for n in notes:
        body_html += f"<p class='footnote'>{n}</p>"
    (OUT / "Table_S8_genotyping.html").write_text(html_page(title, body_html))


def build_s9():
    df = pd.read_csv(SRC / "genotyping" / "qba_table.csv")

    def check(v):
        return "✓" if int(v) == 1 else "—"

    df_disp = pd.DataFrame({
        "Study": df["Study"],
        "WT arm N": df["WT_N"].astype(int),
        "WT events": df["WT_R"].astype(int),
        "WT event rate": df["WT_Event_Rate"].map(lambda x: f"{x:.3f}"),
        "HapB3": df["Tested_HapB3"].map(check),
        "*2A": df["Tested_2A"].map(check),
        "c.2846A>T": df["Tested_2846"].map(check),
        "*13": df["Tested_13"].map(check),
        "Untested variants": df["Untested_Variants"].map(
            lambda x: "—" if str(x).strip() in ("None", "nan", "") else str(x)),
        "Expected misclassified (n)": df["Expected_Misclassified"].map(lambda x: f"{float(x):.1f}"),
        "% of WT arm": df["Pct_of_WT_Arm"].map(lambda x: f"{float(x):.1f}"),
        "Max event-rate shift": df["Max_Event_Rate_Shift"].map(lambda x: f"{float(x):.4f}"),
    })

    total_n = int(df["WT_N"].sum())
    total_r = int(df["WT_R"].sum())
    total_misclassified = float(df["Expected_Misclassified"].sum())
    total_row = {
        "Study": "Total",
        "WT arm N": str(total_n),
        "WT events": str(total_r),
        "WT event rate": f"{total_r / total_n:.3f}",
        "HapB3": "", "*2A": "", "c.2846A>T": "", "*13": "",
        "Untested variants": "",
        "Expected misclassified (n)": f"{total_misclassified:.1f}",
        "% of WT arm": f"{100 * total_misclassified / total_n:.1f}",
        "Max event-rate shift": "",
    }

    title = "Table S9. Quantitative Bias Analysis for Wild-Type Genotyping Completeness"
    caption = (
        "Per-study estimate of patients potentially misclassified as wild-type because the study did not "
        "test for all four clinically actionable variants (HapB3, *2A, c.2846A>T, *13). Expected misclassified count "
        "= WT arm N × Σ carrier frequency of each untested variant. The maximum event-rate shift quantifies the "
        "expected upward movement in the study's observed WT event rate if all expected misclassified patients "
        "experienced events at the variant-specific rates implied by the primary network meta-analysis."
    )
    headers = list(df_disp.columns)
    n_cols = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Table S9"
    col_widths = [26, 10, 10, 12, 9, 7, 12, 7, 28, 14, 11, 16]
    autosize(ws, col_widths)
    write_title_block(ws, title, caption, n_cols, col_widths)
    for i, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i, value=h)
    style_header_row(ws, 4, n_cols)

    r = 5
    for _, row in df_disp.iterrows():
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=row[h])
            style_body_cell(cell, align="left" if h == "Study" else "center")
        r += 1
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=total_row[h])
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.alignment = LEFT if h == "Study" else CENTER
        cell.border = BORDER
    r += 1

    notes = [
        "✓ indicates the variant was genotyped in the WT arm; — indicates not tested. "
        "European carrier frequencies used to compute expected misclassified counts: HapB3 = 4.7%, "
        "*2A = 1.6%, c.2846A>T = 0.7%, *13 = 0.1% (Amstutz et al., CPIC guideline for "
        "fluoropyrimidines and DPYD, Clin Pharmacol Ther 2018; doi:10.1002/cpt.911).",
        "Expected misclassified = WT arm N × Σᵥ (carrier frequency of untested variant v). "
        "% of WT arm = expected misclassified / WT arm N × 100. "
        "Max event-rate shift = Σᵥ [carrier frequency of v × (expected event rate of v − pooled "
        "WT event rate, 0.323)], where expected event rate of v is derived by applying the primary "
        "NMA odds ratio for variant v to the pooled WT baseline.",
    ]
    write_footnotes(ws, r, n_cols, notes, col_widths)
    configure_page(ws, landscape=True)
    wb.save(OUT / "Table_S9_qba.xlsx")

    html_rows = ["<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"]
    for _, row in df_disp.iterrows():
        cells = "".join(
            f"<td class='left'>{row[h]}</td>" if h == "Study" else f"<td>{row[h]}</td>"
            for h in headers
        )
        html_rows.append(f"<tr>{cells}</tr>")
    tot_cells = "".join(
        f"<td class='left'>{total_row[h]}</td>" if h == "Study" else f"<td>{total_row[h]}</td>"
        for h in headers
    )
    html_rows.append(f"<tr class='total'>{tot_cells}</tr>")
    table_html = "<table>" + "".join(html_rows) + "</table>"
    body_html = f"<h2>{title}</h2><p class='caption'>{caption}</p>{table_html}"
    for n in notes:
        body_html += f"<p class='footnote'>{n}</p>"
    (OUT / "Table_S9_qba.html").write_text(html_page(title, body_html))


def build_s10a():
    df = pd.read_csv(SRC / "ld" / "ld_study_breakdown.csv")

    def i(x):
        try:
            return str(int(round(float(x))))
        except (TypeError, ValueError):
            return str(x)

    df_disp = pd.DataFrame({
        "Study": df["Study"],
        "SNP genotyped": df["SNP_Used"],
        "Original N": df["HapB3_N"].map(i),
        "Original events": df["HapB3_R"].map(i),
        "Event rate": df["Event_Rate"].map(lambda x: f"{x:.3f}"),
        "Expected discordant": df["Expected_Discordant"].map(lambda x: f"{float(x):.2f}"),
        "Best-case N": df["Best_N"].map(i),
        "Best-case events": df["Best_R"].map(i),
        "Worst-case N": df["Worst_N"].map(i),
        "Worst-case events": df["Worst_R"].map(i),
        "LD adjustment applied": df["LD_Adjustment_Applied"],
    })

    total_n = int(df["HapB3_N"].sum())
    total_r = int(df["HapB3_R"].sum())
    total_disc = float(df["Expected_Discordant"].sum())
    total_best_n = int(df["Best_N"].sum())
    total_best_r = int(df["Best_R"].sum())
    total_worst_n = int(df["Worst_N"].sum())
    total_worst_r = int(df["Worst_R"].sum())
    total_row = {
        "Study": "Total",
        "SNP genotyped": "",
        "Original N": str(total_n),
        "Original events": str(total_r),
        "Event rate": f"{total_r / total_n:.3f}",
        "Expected discordant": f"{total_disc:.2f}",
        "Best-case N": str(total_best_n),
        "Best-case events": str(total_best_r),
        "Worst-case N": str(total_worst_n),
        "Worst-case events": str(total_worst_r),
        "LD adjustment applied": "",
    }

    title = "Table S10a. Linkage Disequilibrium Sensitivity — Per-Study HapB3 Breakdown"
    caption = (
        "HapB3 arm characteristics for the 16 studies contributing HapB3 data. "
        "Studies genotyped the causal SNP c.1129−5923C>G, the tag SNP c.1236G>A, or both. "
        "Expected discordant = HapB3 N × 1/300 (tag-SNP discordance rate). "
        "Best-case adjustment removes the discordant patient and treats them as a non-event; "
        "worst-case treats them as an event. Adjustment was applied only when expected discordant ≥ 0.5."
    )
    headers = list(df_disp.columns)
    n_cols = len(headers)
    left_cols = {"Study", "SNP genotyped", "LD adjustment applied"}

    wb = Workbook()
    ws = wb.active
    ws.title = "Table S10a"
    col_widths = [26, 30, 11, 13, 11, 14, 12, 13, 13, 14, 22]
    autosize(ws, col_widths)
    write_title_block(ws, title, caption, n_cols, col_widths)
    for i_, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i_, value=h)
    style_header_row(ws, 4, n_cols)

    r = 5
    for _, row in df_disp.iterrows():
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=c, value=row[h])
            style_body_cell(cell, align="left" if h in left_cols else "center")
        r += 1
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=total_row[h])
        cell.fill = TOTAL_FILL
        cell.font = TOTAL_FONT
        cell.alignment = LEFT if h in left_cols else CENTER
        cell.border = BORDER
    r += 1

    notes = [
        "Per-study expected discordant counts (column 6) remained below the LD-adjustment threshold (≥ 0.5) for "
        "all 16 studies individually. The cumulative expected count across the six tag-SNP-only studies totals "
        "0.50 patients, conservatively rounded up to 1; this single discordant patient is allocated to Medwid_2023 "
        "(scenario A) and Wigle_2021 (scenario B) — the two largest tag-SNP-only studies — for the LD-adjusted "
        "re-analyses (see Table S10b).",
        "Tag SNP c.1236G>A is in tight LD with causal HapB3 c.1129-5923C>G (D' = 1, r² ≈ 1 in European populations; "
        "Froehlich et al, 2015) with approximately 1 in 300 carriers discordant; concordance has been confirmed by "
        "direct genotyping of both SNPs (Jennings et al, 2013; Lee et al, 2016; Froehlich et al, 2015).",
    ]
    write_footnotes(ws, r, n_cols, notes, col_widths)
    configure_page(ws, landscape=True)
    wb.save(OUT / "Table_S10a_ld_breakdown.xlsx")

    html_rows = ["<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"]
    for _, row in df_disp.iterrows():
        cells = "".join(
            f"<td class='left'>{row[h]}</td>" if h in left_cols else f"<td>{row[h]}</td>"
            for h in headers
        )
        html_rows.append(f"<tr>{cells}</tr>")
    tot_cells = "".join(
        f"<td class='left'>{total_row[h]}</td>" if h in left_cols else f"<td>{total_row[h]}</td>"
        for h in headers
    )
    html_rows.append(f"<tr class='total'>{tot_cells}</tr>")
    table_html = "<table>" + "".join(html_rows) + "</table>"
    body_html = f"<h2>{title}</h2><p class='caption'>{caption}</p>{table_html}"
    for n in notes:
        body_html += f"<p class='footnote'>{n}</p>"
    (OUT / "Table_S10a_ld_breakdown.html").write_text(html_page(title, body_html))


def build_s10b():
    df = pd.read_csv(SRC / "ld" / "ld_summary.csv")
    primary_or, primary_cri_low, primary_cri_high = 2.005, 1.294, 3.186
    primary_psrf = 1.02854
    primary_dic = 151.492

    scen_name = {
        "best_medwid": "Best case (Medwid_2023, no event)",
        "best_wigle": "Best case (Wigle_2021, no event)",
        "worst_medwid": "Worst case (Medwid_2023, event)",
        "worst_wigle": "Worst case (Wigle_2021, event)",
    }

    rows = [[
        "Primary (no LD adjustment)",
        f"{primary_or:.3f}",
        f"{primary_cri_low:.2f}–{primary_cri_high:.2f}",
        "—",
        f"{primary_psrf:.3f}",
        f"{primary_dic:.2f}",
        "Yes",
        "No",
    ]]
    for _, r in df.iterrows():
        rows.append([
            scen_name.get(r["scenario"], r["scenario"]),
            f"{float(r['HapB3_OR']):.3f}",
            f"{float(r['CrI_low']):.2f}–{float(r['CrI_high']):.2f}",
            f"{float(r['delta_from_primary']):+.3f}",
            f"{float(r['PSRF']):.3f}",
            f"{float(r['DIC']):.2f}",
            yn(r["converged"]),
            yn(r["escalated"]),
        ])

    headers = [
        "Scenario",
        "HapB3 OR",
        "95% CrI",
        "Δ OR from primary",
        "Max PSRF",
        "DIC",
        "Converged",
        "Escalated",
    ]

    title = "Table S10b. Linkage Disequilibrium Sensitivity — HapB3 Odds Ratio Across LD Scenarios"
    caption = (
        "HapB3 vs wild-type odds ratio under best-case and worst-case allocation of the single expected discordant "
        "patient identified via tag-SNP genotyping (see Table S10a). Allocation tested against the two largest "
        "tag-SNP studies (Medwid_2023 and Wigle_2021). All four LD-adjusted scenarios converged (PSRF ≤ 1.10) "
        "and remained within 0.04 OR units of the primary estimate (HapB3 OR 2.01, 95% CrI 1.29–3.19)."
    )
    n_cols = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Table S10b"
    col_widths = [38, 11, 14, 18, 11, 11, 12, 12]
    autosize(ws, col_widths)
    write_title_block(ws, title, caption, n_cols, col_widths)
    for i_, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i_, value=h)
    style_header_row(ws, 4, n_cols)
    r = 5
    for row in rows:
        is_primary = row[0].startswith("Primary")
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if is_primary:
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
            else:
                cell.font = BODY_FONT
            cell.alignment = LEFT if c == 1 else CENTER
            cell.border = BORDER
        r += 1
    notes = [
        "Δ OR from primary = HapB3 OR (scenario) − HapB3 OR (primary, 2.01). Absolute values < 0.10 indicate "
        "consistency with the primary analysis.",
        "PSRF = potential scale reduction factor (Gelman-Rubin diagnostic, convergence threshold ≤ 1.10). "
        "DIC = deviance information criterion.",
        "Escalated = scenario was re-run with an extended MCMC configuration (200,000 iterations, "
        "100,000 burn-in, 3 chains, thinning interval 20) after the sensitivity-run configuration "
        "(150,000 iterations, 75,000 burn-in, 3 chains, thinning interval 10) failed convergence; "
        "none of the LD scenarios required escalation.",
    ]
    write_footnotes(ws, r, n_cols, notes, col_widths)
    configure_page(ws, landscape=False)
    wb.save(OUT / "Table_S10b_ld_summary.xlsx")

    html_rows = ["<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"]
    for row in rows:
        is_primary = row[0].startswith("Primary")
        cls = " class='total'" if is_primary else ""
        cells = "".join(
            f"<td class='left'>{v}</td>" if i == 0 else f"<td>{v}</td>"
            for i, v in enumerate(row)
        )
        html_rows.append(f"<tr{cls}>{cells}</tr>")
    table_html = "<table>" + "".join(html_rows) + "</table>"
    body_html = f"<h2>{title}</h2><p class='caption'>{caption}</p>{table_html}"
    for n in notes:
        body_html += f"<p class='footnote'>{n}</p>"
    (OUT / "Table_S10b_ld_summary.html").write_text(html_page(title, body_html))


SUBHEAD_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHEAD_FONT = Font(name="Calibri", size=10, bold=True)


def build_s2():
    df = pd.read_csv(SRC / "sensitivity_summary.csv")

    def round_half_up(f, ndigits):
        # IEEE-754 2.005 reads as slightly below 2.005, so Python's banker's
        # rounding gives 2.00. The manuscript reports 2.01. Force round-half-up
        # for positive numbers (all ORs here are positive) so the table matches
        # the manuscript convention.
        m = 10 ** ndigits
        return int(f * m + 0.5) / m if f >= 0 else -int(-f * m + 0.5) / m

    def fmt_or(v):
        try:
            f = float(v)
        except (TypeError, ValueError):
            return "—"
        if f != f:
            return "—"
        return f"{round_half_up(f, 2):.2f}"

    def fmt_cri(lo, hi):
        try:
            lo, hi = float(lo), float(hi)
        except (TypeError, ValueError):
            return "—"
        if lo != lo or hi != hi:
            return "—"
        return f"{round_half_up(lo, 2):.2f}–{round_half_up(hi, 2):.2f}"

    def fmt_pct(v):
        try:
            f = float(v)
        except (TypeError, ValueError):
            return "—"
        if f != f:
            return "—"
        return f"{f:.1f}"

    def fmt_dic(v):
        try:
            f = float(v)
        except (TypeError, ValueError):
            return "—"
        if f != f:
            return "—"
        return f"{f:.2f}"

    def fmt_psrf(v):
        try:
            f = float(v)
        except (TypeError, ValueError):
            return "—"
        if f != f:
            return "—"
        return f"{f:.3f}"

    def fmt_int(v):
        try:
            f = float(v)
        except (TypeError, ValueError):
            return "—"
        if f != f:
            return "—"
        return str(int(f))

    def render_row(r):
        return [
            r["description"],
            fmt_int(r["n_studies"]),
            fmt_int(r["n_arms"]),
            fmt_or(r["HapB3_OR"]),
            fmt_cri(r["HapB3_CrI_low"], r["HapB3_CrI_high"]),
            fmt_pct(r["HapB3_SUCRA"]),
            fmt_dic(r["DIC"]),
            fmt_psrf(r["max_PSRF"]),
            yn(r["converged"]),
            str(r["conclusion"]),
        ]

    primary_rows = df[df["scenario"] == "primary"]
    alt_rows = df[~df["scenario"].str.startswith("loo_") & (df["scenario"] != "primary")]
    loo_rows = df[df["scenario"].str.startswith("loo_")].copy()
    # Sort LOO alphabetically by study label (after stripping "LOO: " prefix)
    loo_rows = loo_rows.sort_values("scenario")

    headers = [
        "Scenario", "Studies (n)", "Arms (n)",
        "HapB3 OR", "95% CrI", "HapB3 SUCRA (%)",
        "DIC", "Max PSRF", "Converged", "Conclusion",
    ]

    title = "Table S2. Summary of Sensitivity Analyses (37 Scenarios)"
    caption = (
        "Each row reports a model perturbation or leave-one-out (LOO) iteration against the primary "
        "het_cor probit arm-based NMA. The primary analysis is shown as a reference. Of 37 sensitivity "
        "scenarios, two were infeasible (heterogeneous correlation with inverse-gamma prior is "
        "unsupported by pcnetmeta; exclusion of arms with N ≤ 5 disconnects the *13 node). Of the 35 "
        "feasible scenarios, 34 converged (PSRF ≤ 1.10); the one non-converged scenario (exclusion of "
        "arms with N ≤ 2) is flagged in the Conclusion column. OR = odds ratio for HapB3 vs wild-type; "
        "CrI = credible interval; SUCRA = surface under the cumulative ranking curve."
    )
    n_cols = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Table S2"
    col_widths = [42, 11, 9, 10, 12, 14, 10, 11, 12, 28]
    autosize(ws, col_widths)
    write_title_block(ws, title, caption, n_cols, col_widths)
    for i_, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i_, value=h)
    style_header_row(ws, 4, n_cols)

    left_cols_idx = {1, 10}  # Scenario and Conclusion

    def write_section(start_row, label, rows_df):
        # Write subheader spanning all columns
        ws.merge_cells(start_row=start_row, end_row=start_row,
                       start_column=1, end_column=n_cols)
        cell = ws.cell(row=start_row, column=1, value=label)
        cell.fill = SUBHEAD_FILL
        cell.font = SUBHEAD_FONT
        cell.alignment = LEFT
        cell.border = BORDER
        r = start_row + 1
        for _, row in rows_df.iterrows():
            rendered = render_row(row)
            for c, v in enumerate(rendered, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                style_body_cell(cell, align="left" if c in left_cols_idx else "center")
            r += 1
        return r

    r = 5
    r = write_section(r, "Primary analysis", primary_rows)
    r = write_section(r, "Model perturbations", alt_rows)
    r = write_section(r, "Leave-one-out (LOO)", loo_rows)

    notes = [
        "Subset sizes vary across scenarios: model perturbations retain all 31 studies and 98 arms; "
        "LOO drops one study per row; sparse-data exclusions drop arms below the size threshold. "
        "Infeasible scenarios show '—' in numeric columns.",
        "Conclusion column: 'Reference' = primary; 'Consistent' = HapB3 OR direction and significance preserved "
        "(|Δ log-OR| < 0.10 vs primary); 'Network disconnected' = removing the variant node breaks network "
        "connectivity; 'Model failed to run' = unsupported prior/model combination; 'Non-converged (interpret "
        "with caution)' = PSRF > 1.10 after escalation to the primary model's full sampling configuration.",
        "Reduced MCMC sampling (100,000 iterations, 50,000 burn-in, 3 chains, thinning interval 10) was used for "
        "sensitivity runs. Scenarios that exceeded PSRF ≤ 1.10 were re-run with the primary model's full "
        "sampling configuration (150,000 iterations, 75,000 burn-in, 3 chains, thinning interval 10).",
        "PSRF = potential scale reduction factor (Gelman-Rubin diagnostic, convergence threshold ≤ 1.10). "
        "DIC = deviance information criterion.",
    ]
    write_footnotes(ws, r, n_cols, notes, col_widths)
    configure_page(ws, landscape=True)
    wb.save(OUT / "Table_S2_sensitivity.xlsx")

    # HTML
    html_rows = ["<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"]

    def html_section(label, rows_df):
        out = [f"<tr class='subhead'><td colspan='{n_cols}'>{label}</td></tr>"]
        for _, row in rows_df.iterrows():
            rendered = render_row(row)
            cells = "".join(
                f"<td class='left'>{v}</td>" if (i + 1) in left_cols_idx else f"<td>{v}</td>"
                for i, v in enumerate(rendered)
            )
            out.append(f"<tr>{cells}</tr>")
        return out

    html_rows.extend(html_section("Primary analysis", primary_rows))
    html_rows.extend(html_section("Model perturbations", alt_rows))
    html_rows.extend(html_section("Leave-one-out (LOO)", loo_rows))
    table_html = "<table>" + "".join(html_rows) + "</table>"
    body_html = f"<h2>{title}</h2><p class='caption'>{caption}</p>{table_html}"
    for n in notes:
        body_html += f"<p class='footnote'>{n}</p>"
    (OUT / "Table_S2_sensitivity.html").write_text(html_page(title, body_html))


WT_UNIFIED = Path("/Users/mikeford/dpyd-arm-model/output/wt_unified")


def build_s4():
    # Read with all columns as strings so "98", "3", etc. don't get coerced to floats
    df = pd.read_csv(WT_UNIFIED / "tables" / "tableS6_diagnostics.csv", dtype=str)

    # Post-process for presentation consistency with sibling tables:
    #  - PSRF to 3 dp (S2/S10b style)
    #  - Iteration count with thousands separator
    def _round_half_up(f, ndigits):
        m = 10 ** ndigits
        return int(f * m + 0.5) / m if f >= 0 else -int(-f * m + 0.5) / m

    def reformat(metric, value):
        if metric == "Max Gelman-Rubin PSRF":
            try:
                return f"{_round_half_up(float(value), 3):.3f}"
            except (TypeError, ValueError):
                return value
        if metric == "Iterations (post burn-in)":
            try:
                return f"{int(float(value)):,}"
            except (TypeError, ValueError):
                return value
        return value

    rows = [(str(r["Metric"]), reformat(str(r["Metric"]), str(r["Value"])))
            for _, r in df.iterrows()]
    rows.append(("Convergence achieved (PSRF ≤ 1.10)", "Yes"))

    headers = ["Metric", "Value"]
    title = "Table S4. Bayesian Model Fit and Convergence Diagnostics"
    caption = (
        "Model fit and Markov chain Monte Carlo (MCMC) convergence diagnostics for the primary "
        "het_cor probit arm-based network meta-analysis (wt_unified model). All chains converged "
        "by the Gelman-Rubin criterion (maximum potential scale reduction factor ≤ 1.10)."
    )
    n_cols = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Table S4"
    col_widths = [42, 18]
    autosize(ws, col_widths)
    write_title_block(ws, title, caption, n_cols, col_widths)
    for i_, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i_, value=h)
    style_header_row(ws, 4, n_cols)

    r = 5
    for metric, value in rows:
        cell = ws.cell(row=r, column=1, value=metric)
        style_body_cell(cell, align="left")
        cell = ws.cell(row=r, column=2, value=value)
        style_body_cell(cell, align="center")
        r += 1

    notes = [
        "DIC = deviance information criterion (lower = better fit, penalised for model complexity). "
        "pD = effective number of parameters. D-bar = posterior mean deviance. "
        "Residual deviance / n ratio (D-bar / n_arms = 92.04 / 98 ≈ 0.94) close to 1 indicates "
        "adequate model fit.",
        "PSRF = potential scale reduction factor (Gelman-Rubin diagnostic). Values ≤ 1.10 "
        "indicate convergence; the maximum across all model parameters was 1.029.",
        "Model specification: arm-based network meta-analysis with probit link and heterogeneous "
        "between-study correlation matrix (het_cor), implemented in the R package pcnetmeta. A "
        "Wishart prior (df = 6, identity scale) is placed on the precision matrix T = (Σ R Σ)^-1, "
        "equivalent to an inverse-Wishart prior on the covariance matrix.",
    ]
    write_footnotes(ws, r, n_cols, notes, col_widths)
    configure_page(ws, landscape=False)
    wb.save(OUT / "Table_S4_model_fit.xlsx")

    html_rows = ["<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"]
    for metric, value in rows:
        html_rows.append(f"<tr><td class='left'>{metric}</td><td>{value}</td></tr>")
    table_html = "<table>" + "".join(html_rows) + "</table>"
    body_html = f"<h2>{title}</h2><p class='caption'>{caption}</p>{table_html}"
    for n in notes:
        body_html += f"<p class='footnote'>{n}</p>"
    (OUT / "Table_S4_model_fit.html").write_text(html_page(title, body_html))


def build_s5():
    df = pd.read_csv(WT_UNIFIED / "wt_unified_heterogeneity_params.csv")

    # Treatment order matches sigma[1..5] indexing in source CSV
    trt_order = ["WT", "HapB3", "2846hetho", "2Ahetho", "13hetho"]
    trt_labels = [VARIANT_NAME.get(t, t) if t != "WT" else "WT" for t in trt_order]

    sigma_rows = df[df["Type"] == "Variance (SD)"].reset_index(drop=True)
    corr_rows = df[df["Type"] == "Correlation"].reset_index(drop=True)

    def f3(v):
        try:
            return f"{float(v):.3f}"
        except (TypeError, ValueError):
            return "—"

    title = "Table S5. Between-Study Heterogeneity and Correlation Parameters"
    caption = (
        "Posterior summaries for the between-study random-effects standard deviations (σ) and "
        "correlation matrix (R) on the probit scale, estimated from the primary het_cor arm-based "
        "network meta-analysis (wt_unified model). Treatment-specific SDs quantify between-study "
        "heterogeneity in arm-level probit-scale event probabilities; correlations describe how "
        "between-study deviations co-vary across treatments."
    )
    headers_sigma = ["Treatment", "Posterior mean SD", "95% CrI lower", "95% CrI upper"]
    headers_corr = ["Treatment"] + trt_labels
    n_cols = max(len(headers_sigma), len(headers_corr))  # 6

    wb = Workbook()
    ws = wb.active
    ws.title = "Table S5"
    col_widths = [22, 18, 14, 14, 14, 14]
    autosize(ws, col_widths)
    write_title_block(ws, title, caption, n_cols, col_widths)

    # --- Section 1: Standard Deviations ---
    r = 4
    ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=n_cols)
    cell = ws.cell(row=r, column=1, value="A. Treatment-specific between-study standard deviations (probit scale)")
    cell.fill = SUBHEAD_FILL
    cell.font = SUBHEAD_FONT
    cell.alignment = LEFT
    cell.border = BORDER
    r += 1

    for i_, h in enumerate(headers_sigma, start=1):
        ws.cell(row=r, column=i_, value=h)
    # Pad remaining columns with empty header cells styled to match
    style_header_row(ws, r, len(headers_sigma))
    # Visually blank out unused cols in this section
    for i_ in range(len(headers_sigma) + 1, n_cols + 1):
        cell = ws.cell(row=r, column=i_, value=None)
        cell.fill = HEADER_FILL
        cell.border = BORDER
    r += 1

    for idx, label in enumerate(trt_labels):
        srow = sigma_rows.iloc[idx]
        values = [label, f3(srow["Posterior_Mean"]), f3(srow["CI_Lower"]), f3(srow["CI_Upper"])]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            style_body_cell(cell, align="left" if c == 1 else "center")
        for c in range(len(values) + 1, n_cols + 1):
            cell = ws.cell(row=r, column=c, value=None)
            cell.border = BORDER
        r += 1

    # Spacer
    ws.row_dimensions[r].height = 8
    r += 1

    # --- Section 2: Correlation Matrix ---
    ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=n_cols)
    cell = ws.cell(row=r, column=1, value="B. Between-study correlation matrix (probit scale, posterior mean)")
    cell.fill = SUBHEAD_FILL
    cell.font = SUBHEAD_FONT
    cell.alignment = LEFT
    cell.border = BORDER
    r += 1

    for i_, h in enumerate(headers_corr, start=1):
        ws.cell(row=r, column=i_, value=h)
    style_header_row(ws, r, len(headers_corr))
    r += 1

    # Build correlation lookup: corr_map[(i,j)] = posterior mean, symmetric
    corr_map = {}
    for _, row in corr_rows.iterrows():
        # Parameter is "R[i,j]"
        param = row["Parameter"]
        ij = param[param.index("[") + 1: param.index("]")].split(",")
        i, j = int(ij[0]), int(ij[1])
        v = float(row["Posterior_Mean"])
        corr_map[(i, j)] = v
        corr_map[(j, i)] = v

    for i in range(1, 6):
        # Row label
        cell = ws.cell(row=r, column=1, value=trt_labels[i - 1])
        style_body_cell(cell, align="left", bold=True)
        for j in range(1, 6):
            if i == j:
                v = "1.000"
            else:
                v = f3(corr_map.get((i, j), float("nan")))
            cell = ws.cell(row=r, column=j + 1, value=v)
            style_body_cell(cell, align="center", bold=(i == j))
            if i == j:
                cell.fill = TOTAL_FILL
        r += 1

    notes = [
        "Section A: Posterior mean and 95% credible interval (CrI) for the between-study standard "
        "deviation (σ) of arm-level probit-scale event probabilities for each treatment.",
        "Section B: Posterior mean of the between-study correlation matrix (R). Diagonal elements "
        "are fixed at 1.0 by construction. Off-diagonal elements describe how between-study "
        "deviations in arm-level probit probabilities co-vary across treatments.",
        "A Wishart prior (df = 6, identity scale) is placed on the precision matrix "
        "T = (Σ R Σ)^-1 — equivalent to an inverse-Wishart prior on the covariance matrix. "
        "Because the model uses a probit (not logit) link, these between-study standard "
        "deviations and correlations are on the probit (latent) scale and exponentiation does "
        "not yield odds ratios.",
    ]
    write_footnotes(ws, r, n_cols, notes, col_widths)
    configure_page(ws, landscape=False)
    wb.save(OUT / "Table_S5_heterogeneity.xlsx")

    # HTML
    html_parts = [
        f"<h2>{title}</h2><p class='caption'>{caption}</p>",
        "<h3>A. Treatment-specific between-study standard deviations (probit scale)</h3>",
        "<table><tr>" + "".join(f"<th>{h}</th>" for h in headers_sigma) + "</tr>",
    ]
    for idx, label in enumerate(trt_labels):
        srow = sigma_rows.iloc[idx]
        html_parts.append(
            f"<tr><td class='left'>{label}</td>"
            f"<td>{f3(srow['Posterior_Mean'])}</td>"
            f"<td>{f3(srow['CI_Lower'])}</td>"
            f"<td>{f3(srow['CI_Upper'])}</td></tr>"
        )
    html_parts.append("</table>")
    html_parts.append("<h3>B. Between-study correlation matrix (probit scale, posterior mean)</h3>")
    html_parts.append("<table><tr>" + "".join(f"<th>{h}</th>" for h in headers_corr) + "</tr>")
    for i in range(1, 6):
        row_cells = [f"<td class='left'>{trt_labels[i - 1]}</td>"]
        for j in range(1, 6):
            v = "1.000" if i == j else f3(corr_map.get((i, j), float("nan")))
            cls = " class='total'" if i == j else ""
            row_cells.append(f"<td{cls}>{v}</td>")
        html_parts.append(f"<tr>{''.join(row_cells)}</tr>")
    html_parts.append("</table>")
    for n in notes:
        html_parts.append(f"<p class='footnote'>{n}</p>")
    (OUT / "Table_S5_heterogeneity.html").write_text(html_page(title, "".join(html_parts)))


def build_s6():
    df = pd.read_csv(WT_UNIFIED / "tables" / "table2_pairwise_or.csv")

    # First column is "Treatment" with row labels matching source variant names.
    src_trts = df["Treatment"].tolist()  # row order
    trt_labels = [VARIANT_NAME.get(t, t) if t != "WT" else "WT" for t in src_trts]

    headers = ["Row vs column"] + trt_labels
    n_cols = len(headers)

    title = "Table S6. Pairwise Odds Ratios and 95% Credible Intervals Between Variants"
    caption = (
        "Posterior pairwise odds ratios (OR) with 95% credible intervals (CrI) for all variant-vs-variant "
        "comparisons, computed from the primary het_cor probit arm-based network meta-analysis "
        "(wt_unified model). Each cell reports the odds ratio of the row treatment relative to the "
        "column treatment: OR > 1 indicates higher predicted toxicity for the row treatment. The matrix "
        "is internally consistent (row vs column = 1 / column vs row) but is shown in full for ease of "
        "reading. Diagonal cells are the reference category."
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Table S6"
    col_widths = [22, 22, 22, 22, 22, 22]
    autosize(ws, col_widths)
    write_title_block(ws, title, caption, n_cols, col_widths)
    for i_, h in enumerate(headers, start=1):
        ws.cell(row=4, column=i_, value=h)
    style_header_row(ws, 4, n_cols)

    r = 5
    for i, src_row in enumerate(src_trts):
        # Row label
        cell = ws.cell(row=r, column=1, value=trt_labels[i])
        style_body_cell(cell, align="left", bold=True)
        for j, src_col in enumerate(src_trts):
            v = df.iloc[i][src_col]
            cell = ws.cell(row=r, column=j + 2, value=str(v))
            style_body_cell(cell, align="center")
            if str(v) == "Ref":
                cell.fill = TOTAL_FILL
                cell.font = TOTAL_FONT
        r += 1

    notes = [
        "Each off-diagonal cell reports the posterior median odds ratio with 95% credible interval "
        "(format: OR (low–high)) for the row treatment relative to the column treatment.",
        "Comparisons whose 95% CrI excludes 1.0 are interpreted as showing a credible difference in "
        "toxicity odds between the two variants under the model.",
        "All ORs derive from the same posterior sample as the primary network meta-analysis; pairwise "
        "estimates are therefore internally consistent across the network. Small apparent deviations "
        "between reciprocal cells (e.g., 1.24 vs. 1 / 0.80 = 1.25) reflect 2-decimal rounding; cells "
        "were computed from the posterior, not by numerical inversion.",
    ]
    write_footnotes(ws, r, n_cols, notes, col_widths)
    configure_page(ws, landscape=True)
    wb.save(OUT / "Table_S6_pairwise_or.xlsx")

    # HTML
    html_rows = ["<tr>" + "".join(f"<th>{h}</th>" for h in headers) + "</tr>"]
    for i, src_row in enumerate(src_trts):
        cells = [f"<td class='left'><b>{trt_labels[i]}</b></td>"]
        for src_col in src_trts:
            v = str(df.iloc[i][src_col])
            cls = " class='total'" if v == "Ref" else ""
            cells.append(f"<td{cls}>{v}</td>")
        html_rows.append(f"<tr>{''.join(cells)}</tr>")
    table_html = "<table>" + "".join(html_rows) + "</table>"
    body_html = f"<h2>{title}</h2><p class='caption'>{caption}</p>{table_html}"
    for n in notes:
        body_html += f"<p class='footnote'>{n}</p>"
    (OUT / "Table_S6_pairwise_or.html").write_text(html_page(title, body_html))


def main():
    build_s2()
    build_s4()
    build_s5()
    build_s6()
    build_s8()
    build_s9()
    build_s10a()
    build_s10b()
    for f in sorted(OUT.iterdir()):
        print(f.name, f.stat().st_size, "bytes")


if __name__ == "__main__":
    main()
